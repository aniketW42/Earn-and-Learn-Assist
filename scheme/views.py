from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponseForbidden, HttpResponseRedirect, JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.utils.crypto import get_random_string
from django.contrib.auth import get_user_model
from django.utils.timezone import localdate
from django.db.models import Sum, Count, Q
from django.core.exceptions import ValidationError
from datetime import timedelta, date
from decimal import Decimal
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

from .forms import (SchemeApplicationForm, WorkLogForm, DepartmentForm, 
                   DepartmentInchargeCreationForm, StudentDepartmentAssignmentForm, 
                   BulkStudentAssignmentForm, PaymentRateForm, PaymentReportFilterForm,
                   PaymentCalculationForm, StudentPaymentSearchForm)
from scheme.models import (SchemeApplication, WorkLog, Department, 
                          DepartmentIncharge, StudentDepartmentAssignment,
                          PaymentRate, PaymentCalculation, DepartmentPaymentSummary, PaymentExport)
from users.decorators import role_required, approved_scheme_required
from notifications.models import Notification
from users.models import User
# Create your views here.
def home(request):
    return render(request, 'base.html')
def about(request):
    return render(request, 'scheme/about.html')

@login_required
@role_required('student')
def scheme_registration(request):
    # Check if user already has a scheme application
    existing_application = SchemeApplication.objects.filter(student=request.user).first()
    
    if existing_application:
        # If application exists and needs correction, redirect to update view
        if existing_application.status == 'Correction Required':
            messages.info(request, "Please review and update your application based on the feedback provided.")
            return redirect('update_scheme_application')
        # If application is approved or pending, go to dashboard
        elif existing_application.status in ['Approved', 'Pending']:
            return redirect('student_dashboard')
        # If rejected, allow new application (though this is rare)
        
    if request.method == 'POST':
        form = SchemeApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            scheme_application = form.save(commit=False)
            scheme_application.student = request.user
            scheme_application.save()
            request.user.is_registered = True
            request.user.save()
            messages.success(request, "Your application has been submitted successfully!")
            return redirect('student_dashboard')  
    else:
        form = SchemeApplicationForm()

    return render(request, 'scheme/scheme_registration.html', {'form': form})

@login_required
@role_required('student')
def update_scheme_application(request):
    student = request.user
    application = get_object_or_404(SchemeApplication, student=student)

    # Parse field-specific feedback if available
    field_feedback = {}
    general_comment = ""
    
    if application.comments:
        # Parse the structured feedback
        lines = application.comments.split('\n')
        in_field_section = False
        
        for line in lines:
            line = line.strip()
            if line.startswith('General Comment:'):
                general_comment = line.replace('General Comment:', '').strip()
            elif line == 'Field-specific corrections required:':
                in_field_section = True
            elif in_field_section and line.startswith('•'):
                # Parse field feedback: "• Field Name: Message"
                parts = line[1:].split(':', 1)  # Remove bullet and split on first colon
                if len(parts) == 2:
                    field_name = parts[0].strip().lower().replace(' ', '_')
                    message = parts[1].strip()
                    field_feedback[field_name] = message

    if request.method == "POST":
        form = SchemeApplicationForm(request.POST, request.FILES, instance=application)
        if form.is_valid():
            updated_application = form.save(commit=False)
            updated_application.status = "Pending"
            updated_application.comments = ""  # Clear previous feedback
            updated_application.save()
            messages.success(request, "Your application has been updated and resubmitted successfully!")
            return redirect('student_dashboard')
        else:
            messages.error(request, "Please correct the errors in the form before submitting.")
    else:
        form = SchemeApplicationForm(instance=application)

    context = {
        'form': form,
        'application': application,
        'is_update': True,
        'field_feedback': field_feedback,
        'general_comment': general_comment,
        'feedback_available': bool(application.comments)
    }
    
    return render(request, 'scheme/update_scheme_application.html', context)

@login_required
@role_required('student')
def student_dashboard(request):
    student = request.user  

    # If no application exists, redirect to registration
    if not SchemeApplication.objects.filter(student=student).exists():
        return redirect('scheme_registration')
    
    # Get the student's application
    application = get_object_or_404(SchemeApplication, student=student)
    
    # If application is not approved, show limited dashboard with application status
    if application.status != 'Approved':
        context = {
            'applicant': application,
            'is_approved': False,
            'application_status': application.status,
            'can_submit_work': False,
            'work_logs': [],
            'already_submitted': False,
            'total_hours': 0,
            'last_updated': None,
        }
        return render(request, 'scheme/dashboard.html', context)

    # For approved students, show full dashboard with work log functionality
    today = localdate()  
    already_submitted = WorkLog.objects.filter(student=student, date=today).exists()
    work_logs = WorkLog.objects.filter(student=student).order_by('-date')
    last_work_log = WorkLog.objects.filter(student=student).order_by('-id').first()
    last_updated = {'date':last_work_log.date, 'time':last_work_log.time } if last_work_log else None  # Only date available
    total_hours = WorkLog.objects.filter(student=student, is_verified=True).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0

    # Calculate current month's hours and remaining limit
    # Only count verified hours for monthly limit tracking
    current_month_verified = WorkLog.objects.filter(
        student=student,
        date__year=today.year,
        date__month=today.month,
        is_verified=True,
        is_rejected=False
    ).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    # Also get total submitted hours for this month (for form validation)
    current_month_submitted = WorkLog.objects.filter(
        student=student,
        date__year=today.year,
        date__month=today.month,
        is_rejected=False
    ).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    monthly_limit = 30
    remaining_hours = monthly_limit - current_month_verified
    monthly_limit_reached = remaining_hours <= 0
    is_sunday = today.weekday() == 6  # Sunday is weekday 6

    if request.method == "POST" and not already_submitted and not monthly_limit_reached and not is_sunday:
        form = WorkLogForm(request.POST, user=student)
        if form.is_valid():
            # Double-check to prevent race conditions
            if not WorkLog.objects.filter(student=student, date=today).exists():
                work_log = form.save(commit=False)
                work_log.student = student
                work_log.date = today
                try:
                    work_log.full_clean()  # This will run model validation
                    work_log.save()
                    messages.success(request, "Work log submitted successfully!")
                except ValidationError as e:
                    messages.error(request, f"Validation Error: {', '.join(e.messages)}")
                    form.add_error(None, e)
            else:
                messages.warning(request, "You have already submitted a work log for today.")
            return redirect('student_dashboard')
        else:
            messages.error(request, "Please correct the errors in the form.")  
    else:
        form = WorkLogForm(user=student)

    context = {
        'applicant': application,
        'is_approved': True,
        'application_status': application.status,
        'can_submit_work': True,
        'work_logs': work_logs,
        'already_submitted': already_submitted,
        'form': form,
        'total_hours': total_hours,
        'last_updated': last_updated,
        'current_month_verified': current_month_verified,
        'current_month_submitted': current_month_submitted,
        'monthly_limit': monthly_limit,
        'remaining_hours': remaining_hours,
        'monthly_limit_reached': monthly_limit_reached,
        'is_sunday': is_sunday,
    }
    return render(request, 'scheme/dashboard.html', context)


# Removed redundant submit_work_log view - functionality integrated into student_dashboard

@login_required
@role_required('student')
@approved_scheme_required
def work_summary(request):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    student = request.user
    
    # Get filter parameters
    status_filter = request.GET.get('status', 'all')  # all, verified, pending
    per_page = request.GET.get('per_page', '10')
    
    # Validate per_page parameter
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 15, 25, 50]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    
    # Base queryset - exclude rejected work logs from summary
    work_logs_query = WorkLog.objects.filter(student=student, is_rejected=False).select_related().order_by('-date', '-time')
    
    # Apply status filter
    if status_filter == 'verified':
        work_logs_query = work_logs_query.filter(is_verified=True)
    elif status_filter == 'pending':
        work_logs_query = work_logs_query.filter(is_verified=False)
    
    # Calculate statistics (always based on all non-rejected logs)
    all_logs = WorkLog.objects.filter(student=student, is_rejected=False)
    total_hours = all_logs.aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    verified_hours = all_logs.filter(is_verified=True).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    pending_hours = total_hours - verified_hours
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(work_logs_query, per_page)
    
    try:
        work_logs = paginator.page(page)
    except PageNotAnInteger:
        work_logs = paginator.page(1)
    except EmptyPage:
        work_logs = paginator.page(paginator.num_pages)
    
    # Calculate weekly hours
    weekly_hours = all_logs.filter(
        date__gte=localdate() - timedelta(days=7)
    ).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    context = {
        'work_logs': work_logs,
        'total_hours': total_hours,
        'verified_hours': verified_hours,
        'pending_hours': pending_hours,
        'weekly_hours': weekly_hours,
        'total_entries': work_logs_query.count(),
        'filtered_entries': paginator.count,
        'status_filter': status_filter,
        'per_page': per_page,
        'per_page_options': [5, 10, 15, 25, 50],
    }
    return render(request, 'scheme/work_summary.html', context)


@login_required
@role_required('department_encharge')
def department_dashboard(request):
    # Get the department incharge record
    try:
        dept_incharge = DepartmentIncharge.objects.get(user=request.user)
        department = dept_incharge.department
        
        # Filter work logs for students assigned to this department only
        assigned_students = StudentDepartmentAssignment.objects.filter(
            department=department, is_active=True
        ).values_list('student_id', flat=True)
        
        # Get pending work logs with pagination
        pending_work_logs_list = WorkLog.objects.filter(
            is_verified=False, 
            is_rejected=False,
            student_id__in=assigned_students
        ).order_by('-date', '-time')
        
        pending_paginator = Paginator(pending_work_logs_list, 10)  # 10 items per page
        pending_page = request.GET.get('pending_page', 1)
        pending_work_logs = pending_paginator.get_page(pending_page)

        # Get verified work logs with pagination
        verified_work_logs_list = WorkLog.objects.filter(
            is_verified=True,
            student_id__in=assigned_students
        ).order_by('-date', '-time')
        
        verified_paginator = Paginator(verified_work_logs_list, 15)  # 15 items per page
        verified_page = request.GET.get('verified_page', 1)
        verified_work_logs = verified_paginator.get_page(verified_page)

        # Get student hours summary with pagination
        student_hours_list = WorkLog.objects.filter(
            is_verified=True,
            student_id__in=assigned_students
        ).values('student__username', 'student__first_name', 'student__last_name').annotate(
            total_hours=Sum('hours_worked')
        ).order_by('-total_hours')
        
        summary_paginator = Paginator(student_hours_list, 20)  # 20 items per page
        summary_page = request.GET.get('summary_page', 1)
        student_hours = summary_paginator.get_page(summary_page)

    except DepartmentIncharge.DoesNotExist:
        messages.error(request, "You are not assigned to any department. Please contact the coordinator.")
        return redirect('home')

    context = {
        'department': department,
        'pending_work_logs': pending_work_logs,
        'verified_work_logs': verified_work_logs,
        'student_hours': student_hours,
        'assigned_students_count': assigned_students.count() if assigned_students else 0,
    }
    return render(request, 'scheme/department_dashboard.html', context)

@login_required
@role_required('department_encharge')
def approve_work_log(request, log_id):
    work_log = get_object_or_404(WorkLog, id=log_id)
    work_log.is_verified = True
    work_log.save()
    return HttpResponseRedirect(reverse('department_dashboard'))

@login_required
@role_required('department_encharge')
def reject_work_log(request, log_id):
    work_log = get_object_or_404(WorkLog, id=log_id)
    if request.method == "POST":
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        work_log.is_rejected = True
        work_log.rejection_reason = rejection_reason
        work_log.rejected_by = request.user
        work_log.save()
        messages.success(request, "Work log has been rejected.")
    return HttpResponseRedirect(reverse('department_dashboard'))



@login_required
@role_required('el_coordinator')
def el_coordinator_dashboard(request):
    from .utils import get_paginated_queryset
    
    # Get basic statistics
    pending_applications = SchemeApplication.objects.filter(status="Pending")
    approved_students = SchemeApplication.objects.filter(status="Approved")
    total_hours = WorkLog.objects.aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    # Additional statistics for multi-department system
    departments_count = Department.objects.filter(is_active=True).count()
    unassigned_students_count = User.objects.filter(
        role='student', 
        is_registered=True, 
        studentdepartmentassignment__isnull=True
    ).count()

    # Paginate pending applications for better performance
    pending_page_obj, _, _ = get_paginated_queryset(
        request, pending_applications.order_by('-id'), 
        per_page_default=10
    )

    # Recent work logs with pagination
    recent_work_logs = WorkLog.objects.select_related('student').order_by('-date', '-time')
    work_logs_page_obj, _, _ = get_paginated_queryset(
        request, recent_work_logs, 
        per_page_default=15
    )

    context = {
        "pending_applications": pending_page_obj,
        "approved_students": approved_students,
        "work_logs": work_logs_page_obj,
        "total_hours": total_hours,
        "departments_count": departments_count,
        "unassigned_students_count": unassigned_students_count,
        "pending_applications_count": pending_applications.count(),
        "approved_students_count": approved_students.count(),
    }
    return render(request, "scheme/el_coordinator_dashboard.html", context)

@login_required
@role_required('el_coordinator')
def view_application(request, application_id):
    application = get_object_or_404(SchemeApplication, id=application_id)

    if request.method == "POST":
        action = request.POST.get("action")
        comments = request.POST.get("comments", "").strip()

        try:
            if action == "approve":
                application.status = "Approved"
                messages.success(request, "Application approved successfully.")
            elif action == "reject":
                application.status = "Rejected"
                messages.error(request, "Application has been rejected.")
            elif action == "notify":
                application.status = "Correction Required"
                
                # Handle field-specific feedback
                field_feedback = []
                general_comment = request.POST.get("general_comment", "").strip()
                
                # Collect field-specific feedback
                field_names = [
                    'first_name', 'middle_name', 'last_name', 'address', 'state', 'dob',
                    'annual_income', 'fathers_occupation', 'caste_category', 'department', 'prn_number',
                    'photo', 'application_form', 'income_certificate', 'caste_certificate',
                    'last_year_marksheet', 'domicile_certificate', 'admission_receipt',
                    'aadhar_card', 'bank_passbook', 'caste_validity_certificate'
                ]
                
                for field_name in field_names:
                    field_checked = request.POST.get(f"field_{field_name}")
                    field_message = request.POST.get(f"message_{field_name}", "").strip()
                    
                    if field_checked and field_message:
                        field_display = field_name.replace('_', ' ').title()
                        field_feedback.append(f"• {field_display}: {field_message}")
                
                # Combine general and field-specific feedback
                feedback_parts = []
                if general_comment:
                    feedback_parts.append(f"General Comment: {general_comment}")
                
                if field_feedback:
                    feedback_parts.append("Field-specific corrections required:")
                    feedback_parts.extend(field_feedback)
                
                combined_comments = "\n\n".join(feedback_parts) if feedback_parts else comments
                application.comments = combined_comments
                
                messages.warning(request, "Student has been notified for corrections.")
            else:
                # For approve/reject, use the simple comments
                application.comments = comments
                
            application.save()
            
            # Create notification with error handling
            try:
                notification_message = f'Your Earn & Learn application status has been updated to {application.status}.'
                if application.comments:
                    notification_message += f' Details: {application.comments}'
                
                Notification.objects.create(
                    user=application.student,
                    message=notification_message
                )
            except Exception as e:
                messages.warning(request, "Application updated but notification delivery failed.")
                
        except Exception as e:
            messages.error(request, "An error occurred while processing the application.")
            return render(request, "scheme/view_application.html", {"application": application})

        return redirect("el_coordinator_dashboard")

    context = {"application": application}
    return render(request, "scheme/view_application.html", context)


@login_required
@role_required('el_coordinator')
def registered_students_view(request):
    """View to display all approved students to E&L Coordinator."""
    from .utils import get_paginated_queryset, apply_search_filter, get_filter_context
    
    # Base queryset with work assignment information
    queryset = SchemeApplication.objects.filter(status='Approved').select_related(
        'student', 
        'student__studentdepartmentassignment__department'
    ).order_by('first_name', 'last_name')
    
    # Apply search filter
    search_term = request.GET.get('search', '').strip()
    if search_term:
        queryset = apply_search_filter(
            queryset, search_term,
            ['first_name', 'last_name', 'prn_number', 'department', 'student__username']
        )
    
    # Apply department filter
    department_filter = request.GET.get('department')
    if department_filter:
        queryset = queryset.filter(department__icontains=department_filter)
    
    # Get paginated results
    page_obj, per_page, per_page_options = get_paginated_queryset(request, queryset, per_page_default=15)
    
    # Get filter context
    filter_context = get_filter_context(
        request,
        show_search=True,
        show_export=True,
        per_page_options=per_page_options
    )

    context = {
        'approved_students': page_obj,
        'page_obj': page_obj,
        'per_page': per_page,
        'total_count': queryset.count(),
        **filter_context
    }
    
    return render(request, 'scheme/registered_students.html', context)

@login_required
@role_required('el_coordinator')
def student_worklog_view(request, student_id):
    """
    Display worklogs for a specific student by ID.
    """
    from .utils import get_paginated_queryset, apply_date_range_filter, get_filter_context
    
    student = get_object_or_404(SchemeApplication, id=student_id, status="Approved")

    # Base queryset
    queryset = WorkLog.objects.filter(student=student.student).order_by('-date', '-time')
    
    # Apply status filter
    status_filter = request.GET.get('status')
    if status_filter == 'verified':
        queryset = queryset.filter(is_verified=True)
    elif status_filter == 'pending':
        queryset = queryset.filter(is_verified=False, is_rejected=False)
    elif status_filter == 'rejected':
        queryset = queryset.filter(is_rejected=True)
    
    # Apply date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    queryset = apply_date_range_filter(queryset, date_from, date_to)
    
    # Get paginated results
    page_obj, per_page, per_page_options = get_paginated_queryset(request, queryset, per_page_default=15)
    
    # Calculate statistics (always based on all logs)
    all_logs = WorkLog.objects.filter(student=student.student)
    verified_hours = all_logs.filter(is_verified=True).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    total_hours = all_logs.aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    pending_hours = all_logs.filter(is_verified=False, is_rejected=False).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    # Status options for filter
    status_options = [
        {'value': 'verified', 'label': 'Verified'},
        {'value': 'pending', 'label': 'Pending'},
        {'value': 'rejected', 'label': 'Rejected'},
    ]
    
    # Get filter context
    filter_context = get_filter_context(
        request,
        show_status_filter=True,
        show_date_filter=True,
        show_export=True,
        status_options=status_options,
        per_page_options=per_page_options
    )

    context = {
        'student': student,
        'worklogs': page_obj,
        'page_obj': page_obj,
        'per_page': per_page,
        'verified_hours': verified_hours,
        'total_hours': total_hours,
        'pending_hours': pending_hours,
        'total_count': queryset.count(),
        **filter_context
    }
    
    return render(request, 'scheme/student_worklog.html', context)

@login_required
@role_required('el_coordinator')
def mark_student_completed(request, student_id):
    student = get_object_or_404(SchemeApplication, student__id=student_id)
    
    if student.status == "Completed":
        student.student.is_active = False
        student.student.save()
        messages.success(request, f"{student.student.get_full_name()} has been marked as completed and deactivated.")
    else:
        messages.warning(request, "Only completed students can be deactivated.")

    return redirect('registered_students')

# Department Management Views

@login_required
@role_required('el_coordinator')
def department_list(request):
    """View to list all departments"""
    from .utils import get_paginated_queryset, apply_search_filter, get_filter_context
    
    # Base queryset
    queryset = Department.objects.all().annotate(
        students_count=Count('assigned_students', filter=Q(assigned_students__is_active=True))
    ).order_by('name')
    
    # Apply search filter
    search_term = request.GET.get('search', '').strip()
    if search_term:
        queryset = apply_search_filter(
            queryset, search_term,
            ['name', 'code', 'description']
        )
    
    # Apply status filter
    status_filter = request.GET.get('status')
    if status_filter:
        is_active = status_filter == 'active'
        queryset = queryset.filter(is_active=is_active)
    
    # Get paginated results
    page_obj, per_page, per_page_options = get_paginated_queryset(request, queryset, per_page_default=10)

    # Calculate statistics for all departments (not just current page)
    all_departments = Department.objects.all().annotate(
        students_count=Count('assigned_students', filter=Q(assigned_students__is_active=True))
    )
    
    total_students = sum(dept.students_count for dept in all_departments)
    departments_with_incharge = all_departments.filter(incharge__isnull=False).count()
    departments_without_incharge = all_departments.filter(incharge__isnull=True).count()

    # Status options for filter
    status_options = [
        {'value': 'active', 'label': 'Active'},
        {'value': 'inactive', 'label': 'Inactive'},
    ]
    
    # Get filter context
    filter_context = get_filter_context(
        request,
        show_search=True,
        show_status_filter=True,
        show_export=True,
        status_options=status_options,
        per_page_options=per_page_options
    )

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'total_count': queryset.count(),
        'total_students': total_students,
        'departments_with_incharge': departments_with_incharge,
        'departments_without_incharge': departments_without_incharge,
        **filter_context
    }
    # Ensure departments is set after filter_context to avoid override
    context['departments'] = page_obj
    return render(request, 'scheme/department_list.html', context)

@login_required
@role_required('el_coordinator')
def add_department(request):
    """View to add a new department"""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save(commit=False)
            department.created_by = request.user
            department.save()
            messages.success(request, f"Department '{department.name}' added successfully!")
            return redirect('department_list')
    else:
        form = DepartmentForm()
    
    context = {'form': form}
    return render(request, 'scheme/add_department.html', context)

@login_required
@role_required('el_coordinator')
def edit_department(request, department_id):
    """View to edit an existing department"""
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f"Department '{department.name}' updated successfully!")
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=department)
    
    context = {'form': form, 'department': department}
    return render(request, 'scheme/edit_department.html', context)

@login_required
@role_required('el_coordinator')
def department_detail(request, department_id):
    """View to show department details with students and incharge"""
    department = get_object_or_404(Department, id=department_id)
    assigned_students = StudentDepartmentAssignment.objects.filter(
        department=department, is_active=True
    ).select_related('student').prefetch_related('student__schemeapplication_set')
    
    # Get work logs summary for this department
    student_ids = assigned_students.values_list('student_id', flat=True)
    work_logs = WorkLog.objects.filter(student_id__in=student_ids, is_rejected=False)
    total_hours = work_logs.aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    verified_hours = work_logs.filter(is_verified=True).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0
    
    context = {
        'department': department,
        'assigned_students': assigned_students,
        'total_hours': total_hours,
        'verified_hours': verified_hours,
        'pending_hours': total_hours - verified_hours,
    }
    return render(request, 'scheme/department_detail.html', context)

@login_required
@role_required('el_coordinator')
def create_department_incharge(request):
    """View to create a new department incharge"""
    if request.method == 'POST':
        form = DepartmentInchargeCreationForm(request.POST)
        if form.is_valid():
            # Create the user
            user = form.save(commit=False)
            user.role = 'department_encharge'
            user.save()
            
            # Create department incharge record
            department = form.cleaned_data['department']
            DepartmentIncharge.objects.create(
                user=user,
                department=department,
                assigned_by=request.user
            )
            
            # Send email with credentials
            try:
                send_mail(
                    subject=f'ENL Assist - Department Incharge Credentials for {department.name}',
                    message=f'''
Hello {user.get_full_name()},

You have been assigned as the Department Incharge for {department.name} in the Earn & Learn Scheme Management System.

Your login credentials:
Username: {user.username}
Password: {form.cleaned_data['password1']}

Please login at: http://your-domain.com/users/login/

Best regards,
ENL Coordinator Team
                    ''',
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                messages.success(request, f"Department incharge created and credentials sent to {user.email}")
            except Exception as e:
                messages.warning(request, "Department incharge created but email sending failed. Please share credentials manually.")
            
            return redirect('department_list')
    else:
        form = DepartmentInchargeCreationForm()
    
    context = {'form': form}
    return render(request, 'scheme/create_department_incharge.html', context)

@login_required
@role_required('el_coordinator')
def assign_student_to_department(request):
    """View to assign a student to a department"""
    if request.method == 'POST':
        form = StudentDepartmentAssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.assigned_by = request.user
            assignment.save()
            
            # Create notification for student
            Notification.objects.create(
                user=assignment.student,
                message=f'You have been assigned to {assignment.department.name} department for your Earn & Learn activities.'
            )
            
            messages.success(request, f"Student {assignment.student.get_full_name()} assigned to {assignment.department.name}")
            return redirect('department_list')
    else:
        form = StudentDepartmentAssignmentForm()
    
    context = {'form': form}
    return render(request, 'scheme/assign_student.html', context)

@login_required
@role_required('el_coordinator')
def bulk_assign_students(request):
    """View to assign multiple students to a department"""
    if request.method == 'POST':
        form = BulkStudentAssignmentForm(request.POST)
        if form.is_valid():
            students = form.cleaned_data['students']
            department = form.cleaned_data['department']
            
            assignments = []
            notifications = []
            
            for student in students:
                assignments.append(StudentDepartmentAssignment(
                    student=student,
                    department=department,
                    assigned_by=request.user
                ))
                notifications.append(Notification(
                    user=student,
                    message=f'You have been assigned to {department.name} department for your Earn & Learn activities.'
                ))
            
            # Bulk create
            StudentDepartmentAssignment.objects.bulk_create(assignments)
            Notification.objects.bulk_create(notifications)
            
            messages.success(request, f"{len(students)} students assigned to {department.name} successfully!")
            return redirect('department_list')
    else:
        form = BulkStudentAssignmentForm()
    
    context = {'form': form}
    return render(request, 'scheme/bulk_assign_students.html', context)

@login_required
@role_required('el_coordinator')
def unassign_student(request, assignment_id):
    """View to unassign a student from department"""
    assignment = get_object_or_404(StudentDepartmentAssignment, id=assignment_id)
    
    if request.method == 'POST':
        student_name = assignment.student.get_full_name()
        dept_name = assignment.department.name
        assignment.is_active = False
        assignment.save()
        
        # Notify student
        Notification.objects.create(
            user=assignment.student,
            message=f'You have been unassigned from {dept_name} department. Please contact the coordinator for more information.'
        )
        
        messages.success(request, f"{student_name} unassigned from {dept_name}")
        return redirect('department_detail', department_id=assignment.department.id)
    
    context = {'assignment': assignment}
    return render(request, 'scheme/confirm_unassign.html', context)


# ============================================================================
# PAYMENT MODULE VIEWS
# ============================================================================

@login_required
@role_required('el_coordinator')
def payment_rate_management(request):
    """View for E&L Coordinator to manage payment rates"""
    current_rate = PaymentRate.get_current_rate()
    
    if request.method == 'POST':
        # If updating existing rate
        if current_rate:
            form = PaymentRateForm(request.POST, instance=current_rate)
        else:
            form = PaymentRateForm(request.POST)
            
        if form.is_valid():
            payment_rate = form.save(commit=False)
            payment_rate.set_by = request.user
            payment_rate.save()
            messages.success(request, 'Payment rate updated successfully!')
            return redirect('payment_rate_management')
    else:
        # If editing existing rate
        if current_rate:
            form = PaymentRateForm(instance=current_rate)
        else:
            form = PaymentRateForm()
    
    # Get rate history (all previous rates)
    rate_history = PaymentRate.objects.all().order_by('-updated_at')[:10]
    
    context = {
        'form': form,
        'current_rate': current_rate,
        'rate_history': rate_history,
    } 
    return render(request, 'scheme/payment_rate_management.html', context)


@login_required
@role_required(['el_coordinator'])
def payment_calculation_bulk(request):
    """View for bulk payment calculation - Only EL Coordinator can calculate payments"""
    if request.method == 'POST':
        form = PaymentCalculationForm(request.POST, user=request.user)
        if form.is_valid():
            year = int(form.cleaned_data['year'])
            month = int(form.cleaned_data['month'])
            department = form.cleaned_data['department']
            recalculate = form.cleaned_data['recalculate_existing']
            
            # Validate that a payment rate exists
            current_rate = PaymentRate.get_current_rate()
            if not current_rate:
                messages.error(request, 'No payment rate has been set. Please set a payment rate before calculating payments.')
                return render(request, 'scheme/payment_calculation_bulk.html', {'form': form})
            
            # Determine departments to process
            if department:
                departments = [department]
            elif hasattr(request.user, 'departmentincharge'):
                departments = [request.user.departmentincharge.department]
            else:
                departments = Department.objects.filter(is_active=True)
            
            total_calculated = 0
            total_students = 0
            errors = []
            
            for dept in departments:
                try:
                    # Get students assigned to department with proper filtering
                    students = User.objects.filter(
                        role='student',
                        is_registered=True,
                        studentdepartmentassignment__department=dept,
                        studentdepartmentassignment__is_active=True
                    ).distinct()  # Use distinct to avoid duplicates
                    
                    dept_calculated = 0
                    
                    for student in students:
                        try:
                            # Check if payment record already exists
                            payment_date = date(year, month, 1)
                            existing_record = PaymentCalculation.objects.filter(
                                student=student,
                                calculation_month=payment_date
                            ).first()
                            
                            if existing_record and not recalculate:
                                continue
                            
                            # Check if student has any verified work logs for the month
                            has_work_logs = WorkLog.objects.filter(
                                student=student,
                                date__year=year,
                                date__month=month,
                                is_verified=True,
                                is_rejected=False
                            ).exists()
                            
                            if not has_work_logs:
                                continue  # Skip students with no verified work logs
                            
                            payment_calculation = PaymentCalculation.calculate_for_student_month(student, year, month)
                            if payment_calculation:
                                payment_calculation.calculated_by = request.user
                                payment_calculation.save()
                                dept_calculated += 1
                                total_calculated += 1
                                
                        except Exception as e:
                            error_msg = f"Error calculating payment for {student.get_full_name()}: {str(e)}"
                            errors.append(error_msg)
                            continue
                    
                    # Generate department summary only if calculations were made
                    if dept_calculated > 0:
                        try:
                            DepartmentPaymentSummary.generate_for_department_month(dept, year, month)
                        except Exception as e:
                            error_msg = f"Error generating summary for {dept.name}: {str(e)}"
                            errors.append(error_msg)
                    
                    total_students += dept_calculated
                    
                except Exception as e:
                    error_msg = f"Error processing department {dept.name}: {str(e)}"
                    errors.append(error_msg)
                    continue
            
            # Provide feedback to user
            if total_calculated > 0:
                messages.success(request, f'Payment calculations completed for {total_calculated} students across {len(departments)} departments!')
            else:
                messages.info(request, 'No new payment calculations were needed. Students may not have verified work logs for the selected period.')
            
            # Display any errors that occurred
            if errors:
                for error in errors[:5]:  # Limit to first 5 errors to avoid overwhelming the user
                    messages.warning(request, error)
                if len(errors) > 5:
                    messages.warning(request, f'... and {len(errors) - 5} more errors occurred.')
            
            return redirect('payment_reports')
    else:
        form = PaymentCalculationForm(user=request.user)
    
    # Get current payment rate for display
    current_rate = PaymentRate.get_current_rate()
    
    context = {
        'form': form,
        'current_rate': current_rate
    }
    return render(request, 'scheme/payment_calculation_bulk.html', context)


@login_required
@role_required(['el_coordinator', 'department_encharge'])
def payment_reports(request):
    """View for payment reports with filtering"""
    filter_form = PaymentReportFilterForm(request.GET, user=request.user)
    
    # Default to current month/year
    year = int(request.GET.get('year', date.today().year))
    month = int(request.GET.get('month', date.today().month))
    department_id = request.GET.get('department')
    student_id = request.GET.get('student')
    
    # Base queryset
    payment_calculations = PaymentCalculation.objects.filter(
        calculation_month__year=year,
        calculation_month__month=month
    )
    
    # Apply filters based on user role
    if hasattr(request.user, 'departmentincharge'):
        # Department incharge can only see their department
        payment_calculations = payment_calculations.filter(
            department=request.user.departmentincharge.department
        )
    elif department_id:
        payment_calculations = payment_calculations.filter(department_id=department_id)
    
    if student_id:
        payment_calculations = payment_calculations.filter(student_id=student_id)
    
    # Get department summaries
    if hasattr(request.user, 'departmentincharge'):
        dept_summaries = DepartmentPaymentSummary.objects.filter(
            calculation_month__year=year,
            calculation_month__month=month,
            department=request.user.departmentincharge.department
        )
    else:
        dept_summaries = DepartmentPaymentSummary.objects.filter(
            calculation_month__year=year,
            calculation_month__month=month
        )
        if department_id:
            dept_summaries = dept_summaries.filter(department_id=department_id)
    
    # Calculate totals
    total_amount = payment_calculations.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_students = payment_calculations.count()
    total_hours = payment_calculations.aggregate(Sum('total_hours'))['total_hours__sum'] or 0
    
    # Pagination
    paginator = Paginator(payment_calculations.order_by('-created_at'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'filter_form': filter_form,
        'payment_records': page_obj,
        'dept_summaries': dept_summaries,
        'total_amount': total_amount,
        'total_students': total_students,
        'total_hours': total_hours,
        'selected_month': calendar.month_name[month],
        'selected_year': year,
    }
    return render(request, 'scheme/payment_reports.html', context)


@login_required
@role_required(['el_coordinator', 'department_encharge'])
def payment_calculation_detail(request, record_id):
    """View for payment calculation details"""
    payment_calculation = get_object_or_404(PaymentCalculation, id=record_id)
    
    # Check permissions
    if hasattr(request.user, 'departmentincharge'):
        if payment_calculation.department != request.user.departmentincharge.department:
            return HttpResponseForbidden("You don't have permission to view this payment calculation.")
    
    # Get related work logs
    work_logs = WorkLog.objects.filter(
        student=payment_calculation.student,
        date__year=payment_calculation.calculation_month.year,
        date__month=payment_calculation.calculation_month.month,
        is_verified=True,
        is_rejected=False
    ).order_by('date')
    
    context = {
        'payment_calculation': payment_calculation,
        'work_logs': work_logs,
    }
    return render(request, 'scheme/payment_calculation_detail.html', context)


@login_required
@role_required('student')
@approved_scheme_required
def student_payment_dashboard(request):
    """Dashboard for students to view their payment information"""
    try:
        # Get current month's payment info
        current_date = date.today()
        current_month_record = PaymentCalculation.objects.filter(
            student=request.user,
            calculation_month__year=current_date.year,
            calculation_month__month=current_date.month
        ).first()
        
        # Get current payment rate
        current_rate = PaymentRate.get_current_rate()
        
        # Get current month's verified hours using database aggregation
        current_month_hours_result = WorkLog.objects.filter(
            student=request.user,
            date__year=current_date.year,
            date__month=current_date.month,
            is_verified=True,
            is_rejected=False
        ).aggregate(Sum('hours_worked'))
        
        current_month_hours = current_month_hours_result['hours_worked__sum'] or 0
        
        # Calculate potential earnings for current month
        potential_earnings = 0
        if current_rate and current_month_hours:
            from decimal import Decimal
            potential_earnings = Decimal(str(current_month_hours)) * current_rate.rate_per_hour
        
        # Get payment history with search
        search_form = StudentPaymentSearchForm(request.GET)
        payment_history = PaymentCalculation.objects.filter(student=request.user)
        
        if search_form.is_valid():
            year = search_form.cleaned_data.get('year')
            month = search_form.cleaned_data.get('month')
            
            if year:
                payment_history = payment_history.filter(calculation_month__year=year)
            if month:
                payment_history = payment_history.filter(calculation_month__month=month)
        
        payment_history = payment_history.order_by('-calculation_month')[:12]  # Last 12 months
        
        # Calculate totals using database aggregation
        total_earned_result = PaymentCalculation.objects.filter(
            student=request.user
        ).aggregate(Sum('total_amount'))
        
        total_earned = total_earned_result['total_amount__sum'] or 0
        
        context = {
            'current_month_record': current_month_record,
            'current_rate': current_rate,
            'current_month_hours': current_month_hours,
            'potential_earnings': potential_earnings,
            'payment_history': payment_history,
            'total_earned': total_earned,
            'search_form': search_form,
        }
        return render(request, 'scheme/student_payment_dashboard.html', context)
        
    except Exception as e:
        # Log the error and show a user-friendly message
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in student payment dashboard for user {request.user.username}: {str(e)}")
        
        messages.error(request, 'An error occurred while loading your payment dashboard. Please try again or contact support.')
        return redirect('student_dashboard')


@login_required
@role_required(['el_coordinator', 'department_encharge'])
def export_payment_report(request):
    """Export payment reports to Excel"""
    year = int(request.GET.get('year', date.today().year))
    month = int(request.GET.get('month', date.today().month))
    export_type = request.GET.get('type', 'department')
    department_id = request.GET.get('department')
    student_id = request.GET.get('student')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if export_type == 'student' and student_id:
        # Individual student report
        student = get_object_or_404(User, id=student_id)
        payment_calculation = PaymentCalculation.objects.filter(
            student=student,
            calculation_month__year=year,
            calculation_month__month=month
        ).first()
        
        ws.title = f"{student.get_full_name()} - {calendar.month_name[month]} {year}"
        
        # Headers
        headers = ['Date', 'Hours', 'Description', 'Verified', 'Rate/Hour', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, col=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Get work logs
        work_logs = WorkLog.objects.filter(
            student=student,
            date__year=year,
            date__month=month
        ).order_by('date')
        
        row = 2
        total_amount = 0
        rate_per_hour = payment_calculation.rate_per_hour if payment_calculation else (PaymentRate.get_current_rate().rate_per_hour if PaymentRate.get_current_rate() else 0)
        
        for log in work_logs:
            ws.cell(row=row, col=1, value=log.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, col=2, value=log.hours_worked)
            ws.cell(row=row, col=3, value=log.description[:50] + '...' if len(log.description) > 50 else log.description)
            ws.cell(row=row, col=4, value='Yes' if log.is_verified else 'No')
            ws.cell(row=row, col=5, value=float(rate_per_hour) if log.is_verified else 0)
            amount = float(rate_per_hour * log.hours_worked) if log.is_verified else 0
            ws.cell(row=row, col=6, value=amount)
            if log.is_verified:
                total_amount += amount
            row += 1
        
        # Add total row
        ws.cell(row=row + 1, col=5, value="Total:").font = Font(bold=True)
        ws.cell(row=row + 1, col=6, value=total_amount).font = Font(bold=True)
        
        filename = f"payment_report_{student.get_full_name().replace(' ', '_')}_{year}_{month:02d}.xlsx"
        
    elif export_type == 'department':
        # Department report
        if department_id:
            department = get_object_or_404(Department, id=department_id)
            departments = [department]
            ws.title = f"{department.name} - {calendar.month_name[month]} {year}"
        elif hasattr(request.user, 'departmentincharge'):
            departments = [request.user.departmentincharge.department]
            ws.title = f"{departments[0].name} - {calendar.month_name[month]} {year}"
        else:
            departments = Department.objects.filter(is_active=True)
            ws.title = f"All Departments - {calendar.month_name[month]} {year}"
        
        # Headers
        headers = ['Student Name', 'PRN', 'Department', 'Total Hours', 'Rate/Hour', 'Total Amount', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, col=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        grand_total = 0
        
        for dept in departments:
            payment_calculations = PaymentCalculation.objects.filter(
                department=dept,
                calculation_month__year=year,
                calculation_month__month=month
            ).order_by('student__first_name')
            
            for record in payment_calculations:
                ws.cell(row=row, col=1, value=record.student.get_full_name())
                # Get PRN from scheme application
                try:
                    prn = record.student.schemeapplication.prn_number
                except:
                    prn = 'N/A'
                ws.cell(row=row, col=2, value=prn)
                ws.cell(row=row, col=3, value=record.department.name)
                ws.cell(row=row, col=4, value=float(record.total_hours))
                ws.cell(row=row, col=5, value=float(record.rate_per_hour))
                ws.cell(row=row, col=6, value=float(record.total_amount))
                ws.cell(row=row, col=7, value=record.get_status_display())
                grand_total += float(record.total_amount)
                row += 1
        
        # Add total row
        ws.cell(row=row + 1, col=5, value="Grand Total:").font = Font(bold=True)
        ws.cell(row=row + 1, col=6, value=grand_total).font = Font(bold=True)
        
        filename = f"payment_report_departments_{year}_{month:02d}.xlsx"
    
    else:
        # All departments summary
        ws.title = f"Payment Summary - {calendar.month_name[month]} {year}"
        
        # Headers
        headers = ['Department', 'Total Students', 'Total Hours', 'Total Amount', 'Avg Hours/Student']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, col=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        summaries = DepartmentPaymentSummary.objects.filter(
            calculation_month__year=year,
            calculation_month__month=month
        ).order_by('department__name')
        
        row = 2
        grand_total_amount = 0
        grand_total_students = 0
        grand_total_hours = 0
        
        for summary in summaries:
            ws.cell(row=row, col=1, value=summary.department.name)
            ws.cell(row=row, col=2, value=summary.total_students)
            ws.cell(row=row, col=3, value=float(summary.total_hours))
            ws.cell(row=row, col=4, value=float(summary.total_amount))
            ws.cell(row=row, col=5, value=float(summary.average_hours_per_student))
            
            grand_total_amount += float(summary.total_amount)
            grand_total_students += summary.total_students
            grand_total_hours += float(summary.total_hours)
            row += 1
        
        # Add totals
        ws.cell(row=row + 1, col=1, value="TOTALS:").font = Font(bold=True)
        ws.cell(row=row + 1, col=2, value=grand_total_students).font = Font(bold=True)
        ws.cell(row=row + 1, col=3, value=grand_total_hours).font = Font(bold=True)
        ws.cell(row=row + 1, col=4, value=grand_total_amount).font = Font(bold=True)
        ws.cell(row=row + 1, col=5, value=grand_total_hours/grand_total_students if grand_total_students > 0 else 0).font = Font(bold=True)
        
        filename = f"payment_summary_all_departments_{year}_{month:02d}.xlsx"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create export record
    PaymentExport.objects.create(
        export_type=export_type,
        export_month=date(year, month, 1),
        department_id=department_id,
        student_id=student_id,
        file_name=filename,
        file_path=f"exports/{filename}",
        exported_by=request.user
    )
    
    # Return file
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required(['el_coordinator', 'department_encharge'])
def department_payment_budget(request):
    """View for department payment budget overview with filtering options"""
    # Get filter parameters
    view_type = request.GET.get('view_type', 'yearly')  # yearly, monthly, custom
    current_year = int(request.GET.get('year', date.today().year))
    current_month = int(request.GET.get('month', date.today().month))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get department budget data
    departments = Department.objects.filter(is_active=True)
    
    if hasattr(request.user, 'departmentincharge'):
        departments = departments.filter(id=request.user.departmentincharge.department.id)
    
    budget_data = []
    total_budget = 0
    
    for dept in departments:
        if view_type == 'monthly':
            # Get data for specific month
            monthly_summaries = DepartmentPaymentSummary.objects.filter(
                department=dept,
                calculation_month__year=current_year,
                calculation_month__month=current_month
            )
            month_total = sum(float(s.total_amount) for s in monthly_summaries)
            monthly_totals = [0] * 12
            monthly_totals[current_month - 1] = month_total
            year_total = month_total
            
        elif view_type == 'custom' and start_date and end_date:
            # Get data for custom date range
            from datetime import datetime
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            monthly_summaries = DepartmentPaymentSummary.objects.filter(
                department=dept,
                calculation_month__gte=start_dt,
                calculation_month__lte=end_dt
            ).order_by('calculation_month')
            
            monthly_totals = [0] * 12
            year_total = 0
            for summary in monthly_summaries:
                month_index = summary.calculation_month.month - 1
                monthly_totals[month_index] = float(summary.total_amount)
                year_total += float(summary.total_amount)
                
        else:
            # Default yearly view
            monthly_summaries = DepartmentPaymentSummary.objects.filter(
                department=dept,
                calculation_month__year=current_year
            ).order_by('calculation_month')
            
            monthly_totals = [0] * 12  # Initialize 12 months
            year_total = 0
            
            for summary in monthly_summaries:
                month_index = summary.calculation_month.month - 1
                monthly_totals[month_index] = float(summary.total_amount)
                year_total += float(summary.total_amount)
        
        # Calculate quarterly totals
        q1_total = sum(monthly_totals[0:3])  # Jan-Mar
        q2_total = sum(monthly_totals[3:6])  # Apr-Jun
        q3_total = sum(monthly_totals[6:9])  # Jul-Sep
        q4_total = sum(monthly_totals[9:12]) # Oct-Dec
        
        # Get current month total for easy template access
        current_month_total = monthly_totals[current_month - 1] if view_type == 'monthly' else year_total
        
        # Calculate student count based on students with approved applications assigned to this department
        # Get students who are assigned to this department and have approved scheme applications
        approved_student_ids = dept.assigned_students.filter(
            is_active=True,
            student__schemeapplication__status='Approved'
        ).values_list('student_id', flat=True)
        
        if view_type == 'monthly':
            # Count approved students who have payment calculations for the specific month
            student_count = PaymentCalculation.objects.filter(
                student_id__in=approved_student_ids,
                department=dept,
                calculation_month__year=current_year,
                calculation_month__month=current_month
            ).values('student').distinct().count()
        elif view_type == 'custom' and start_date and end_date:
            # Count approved students who have payment calculations in the custom date range
            student_count = PaymentCalculation.objects.filter(
                student_id__in=approved_student_ids,
                department=dept,
                calculation_month__gte=start_dt,
                calculation_month__lte=end_dt
            ).values('student').distinct().count()
        else:
            # For yearly view, just count all approved students assigned to department
            student_count = len(approved_student_ids)
        
        budget_data.append({
            'department': dept,
            'monthly_totals': monthly_totals,
            'quarterly_totals': [q1_total, q2_total, q3_total, q4_total],
            'year_total': year_total,
            'current_month_total': current_month_total,
            'student_count': student_count
        })
        
        total_budget += year_total
    
    # Calculate totals for footer
    total_students = sum(dept_data['student_count'] for dept_data in budget_data)
    total_quarterly = [0, 0, 0, 0]
    for dept_data in budget_data:
        for i in range(4):
            total_quarterly[i] += dept_data['quarterly_totals'][i]
    
    # Get monthly labels
    months = [calendar.month_abbr[i] for i in range(1, 13)]
    
    context = {
        'budget_data': budget_data,
        'total_budget': total_budget,
        'total_students': total_students,
        'total_quarterly': total_quarterly,
        'current_year': current_year,
        'current_month': current_month,
        'view_type': view_type,
        'start_date': start_date,
        'end_date': end_date,
        'months': months,
    }
    return render(request, 'scheme/department_payment_budget.html', context)


@login_required
def scheme_approval_required_view(request):
    """View shown when student needs scheme approval to access features"""
    return render(request, 'scheme/scheme_approval_required.html')  